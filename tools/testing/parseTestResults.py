#!/usr/bin/env python3
"""
Parse completed test results from testing_results directories and generate Excel reports.
Creates version-specific worksheets in a master Excel file for tracking test execution.

Usage:
    python3 parseTestResults.py v3.0.0                    # Parse v3.0.0 results
    python3 parseTestResults.py v3.0.0 --output custom.xlsx  # Custom output file

Requirements:
    pip install openpyxl
"""

import os
import sys
import re
import argparse
from pathlib import Path

# Import Excel dependencies
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl library is required for this script.")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def build_testing_results_path(version):
    """Build testing results path from version string."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    scopy_root = os.path.dirname(os.path.dirname(script_dir))  # tools/testing/ → scopy/
    return os.path.join(scopy_root, "testing_results", f"testing_results_{version}")


def get_output_excel_path(custom_path=None):
    """Get Excel output path (default: testing_results/scopy_test_results.xlsx)."""
    if custom_path:
        return custom_path

    script_dir = os.path.dirname(os.path.abspath(__file__))
    scopy_root = os.path.dirname(os.path.dirname(script_dir))  # tools/testing/ → scopy/
    return os.path.join(scopy_root, "testing_results", "scopy_test_results.xlsx")


def scan_testing_results_files(results_dir):
    """Find all RST files in testing results directory."""
    rst_files = []

    for root, dirs, files in os.walk(results_dir):
        for file in files:
            if file.endswith('.rst'):
                file_path = os.path.join(root, file)
                relative_path = os.path.relpath(file_path, results_dir)
                component = extract_component_from_testing_path(relative_path)

                rst_files.append({
                    'file_path': file_path,
                    'relative_path': relative_path,
                    'component': component,
                    'filename': file
                })

    return rst_files


def extract_component_from_testing_path(relative_path):
    """Extract component from testing results file path."""
    parts = relative_path.split('/')

    if parts[0] == 'plugins' and len(parts) >= 2:
        return parts[1]  # plugins/adc/... → 'adc'
    elif parts[0] == 'general':
        if len(parts) >= 2 and parts[1] == 'core':
            return 'core'
        elif 'preferences' in parts[-1]:
            return 'preferences'
        elif 'package_manager' in parts[-1]:
            return 'package_manager'
        elif 'scripting' in parts[-1]:
            return 'scripting'
        elif 'instrument_detaching' in parts[-1]:
            return 'instrument_detaching'
        else:
            return 'general'
    else:
        return 'other'


def parse_rst_structure(content):
    """Parse RST into header section + individual test sections."""
    # Find test boundaries using pattern: "Test N - " followed by underline
    test_pattern = r'^(Test \d+[^\n]*\n[\^=-]+\n)'

    # Split content at test boundaries
    parts = re.split(test_pattern, content, flags=re.MULTILINE)

    if len(parts) == 1:
        # No tests found
        return {'header': content, 'tests': []}

    header = parts[0]  # Everything before first test
    tests = []

    # Process test sections (parts[1], parts[2], parts[3], parts[4], ...)
    for i in range(1, len(parts), 2):
        if i + 1 < len(parts):
            heading = parts[i]
            content_block = parts[i + 1]

            # Extract UID/RBP from content
            uid = extract_uid_from_content(content_block)
            rbp = extract_rbp_from_content(content_block)

            # Extract result metadata
            result_data = extract_result_metadata(content_block)

            tests.append({
                'heading': heading,
                'content': content_block,
                'uid': uid,
                'rbp': rbp,
                'result': result_data['result'],
                'tested_os': result_data['tested_os'],
                'comments': result_data['comments']
            })

    return {'header': header, 'tests': tests}


def extract_uid_from_content(content):
    """Extract UID from test content block."""
    uid_pattern = r'\*\*UID:\*\*\s+([A-Z_][A-Z0-9_.]*)'
    match = re.search(uid_pattern, content)
    return match.group(1) if match else 'MISSING'


def extract_rbp_from_content(content):
    """Extract RBP from test content block."""
    rbp_pattern = r'\*\*RBP:\*\*\s+(P[0-3])'
    match = re.search(rbp_pattern, content)
    return match.group(1) if match else 'MISSING'


def extract_result_metadata(content):
    """Extract result fields from test content."""
    result_data = {
        'result': None,
        'tested_os': None,
        'comments': None
    }

    # Parse **Result:** PASS/FAIL (treat "PASS/FAIL" as SKIPPED)
    result_pattern = r'\*\*Result:\*\*\s+([^\n]+)'
    result_match = re.search(result_pattern, content)
    if result_match:
        result_value = result_match.group(1).strip()
        if result_value == 'PASS/FAIL':
            result_data['result'] = None  # Will become SKIPPED
        elif result_value in ['PASS', 'FAIL']:
            result_data['result'] = result_value
        else:
            result_data['result'] = None  # Unknown format, treat as SKIPPED

    # Parse **Tested OS:** value
    os_pattern = r'\*\*Tested OS:\*\*\s+([^\n]+)'
    os_match = re.search(os_pattern, content)
    if os_match:
        result_data['tested_os'] = os_match.group(1).strip()

    # Parse **Comments:** value
    comments_pattern = r'\*\*Comments:\*\*\s+([^\n]+)'
    comments_match = re.search(comments_pattern, content)
    if comments_match:
        result_data['comments'] = comments_match.group(1).strip()

    return result_data


def normalize_result_status(raw_status):
    """Convert PASS/FAIL to PASSED/FAILED, handle missing."""
    if raw_status == 'PASS':
        return 'PASSED'
    elif raw_status == 'FAIL':
        return 'FAILED'
    else:
        return 'SKIPPED'


def parse_test_results_from_rst(file_path):
    """Parse RST file and extract all test results."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        structure = parse_rst_structure(content)
        return structure['tests']

    except Exception as e:
        print(f"Warning: Error parsing {file_path}: {e}")
        return []


def process_all_test_data(results_dir):
    """Main processing function - parse all test data from directory."""
    all_test_data = []

    rst_files = scan_testing_results_files(results_dir)

    for rst_file in rst_files:
        # Skip index files - they don't contain tests
        if 'index.rst' in rst_file['filename']:
            continue

        tests = parse_test_results_from_rst(rst_file['file_path'])

        for test in tests:
            test_data = {
                'uid': test['uid'],
                'component': rst_file['component'],
                'rbp': test['rbp'],
                'result': normalize_result_status(test['result']),
                'tested_os': test['tested_os'] or 'N/A',
                'comments': test['comments'] or 'N/A',
                'file': rst_file['filename']
            }
            all_test_data.append(test_data)

    return all_test_data


def validate_and_fill_missing_data(test_data):
    """Ensure all fields have values (N/A for missing)."""
    for test in test_data:
        for key, value in test.items():
            if value is None or value == '' or value == 'MISSING':
                if key == 'result':
                    test[key] = 'SKIPPED'
                else:
                    test[key] = 'N/A'

    return test_data


def setup_worksheet_formatting(worksheet):
    """Apply Excel formatting (headers, column widths, etc.)."""
    # Format header row
    for col in range(1, 8):  # A through G
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Auto-size columns
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].auto_size = True


def generate_excel_report(test_data, excel_path, version):
    """Generate or update Excel file with version worksheet."""

    # Load existing workbook or create new one
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Create or get worksheet for this version
    if version in wb.sheetnames:
        ws = wb[version]
        # Clear existing content
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(version)

    # Write headers
    headers = ['Test UID', 'Component', 'RBP', 'Result', 'Tested OS', 'Comments', 'File']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write test data
    for row_idx, test in enumerate(test_data, 2):  # Start from row 2
        ws.cell(row=row_idx, column=1, value=test['uid'])
        ws.cell(row=row_idx, column=2, value=test['component'])
        ws.cell(row=row_idx, column=3, value=test['rbp'])
        ws.cell(row=row_idx, column=4, value=test['result'])
        ws.cell(row=row_idx, column=5, value=test['tested_os'])
        ws.cell(row=row_idx, column=6, value=test['comments'])
        ws.cell(row=row_idx, column=7, value=test['file'])

    # Apply formatting
    setup_worksheet_formatting(ws)

    # Save workbook
    wb.save(excel_path)
    return True


def get_test_summary_stats(test_data):
    """Generate summary statistics for reporting."""
    total_tests = len(test_data)

    # Count by status
    status_counts = {}
    for test in test_data:
        status = test['result']
        status_counts[status] = status_counts.get(status, 0) + 1

    # Count by component
    component_counts = {}
    for test in test_data:
        component = test['component']
        component_counts[component] = component_counts.get(component, 0) + 1

    return {
        'total': total_tests,
        'by_status': status_counts,
        'by_component': component_counts
    }


def print_summary_report(test_data, excel_path, version, stats):
    """Print detailed summary report."""
    print(f"\n=== Test Results Parsing Summary ===")
    print(f"Version: {version}")
    print(f"Excel Report: {excel_path}")
    print(f"Worksheet: \"{version}\" ({stats['total']} tests)")

    # Component breakdown
    print(f"\nProcessing by component:")
    for component, count in stats['by_component'].items():
        # Calculate component-specific status breakdown
        comp_tests = [t for t in test_data if t['component'] == component]
        comp_stats = get_test_summary_stats(comp_tests)

        passed = comp_stats['by_status'].get('PASSED', 0)
        failed = comp_stats['by_status'].get('FAILED', 0)
        skipped = comp_stats['by_status'].get('SKIPPED', 0)

        print(f"├── {component.title()} Plugin: {count} tests ({passed} PASSED, {failed} FAILED, {skipped} SKIPPED)")

    # Overall summary
    print(f"\n=== Overall Summary ===")
    print(f"Total Tests: {stats['total']}")

    for status in ['PASSED', 'FAILED', 'SKIPPED']:
        count = stats['by_status'].get(status, 0)
        percentage = (count / stats['total'] * 100) if stats['total'] > 0 else 0
        print(f"├── {status}: {count} ({percentage:.0f}%)")

    print(f"\n✓ Test results parsing completed successfully!")


def main():
    parser = argparse.ArgumentParser(
        description='Parse Scopy test results and generate Excel report',
        epilog='Requires: openpyxl (pip install openpyxl)',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('version', help='Version to parse (e.g., v3.0.0)')
    parser.add_argument('--output', help='Custom output Excel file path')

    args = parser.parse_args()

    # Build paths
    results_dir = build_testing_results_path(args.version)
    excel_path = get_output_excel_path(args.output)

    # Validate input directory exists
    if not os.path.exists(results_dir):
        print(f"Error: Testing results directory not found: {results_dir}")
        print(f"Make sure you've run setup_test_environment.py for version {args.version}")
        sys.exit(1)

    # Create output directory if needed
    output_dir = os.path.dirname(excel_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"Parsing test results for version: {args.version}")
    print(f"Source directory: {results_dir}")

    # Process all test data
    test_data = process_all_test_data(results_dir)
    test_data = validate_and_fill_missing_data(test_data)

    if not test_data:
        print("Warning: No test data found in the specified directory.")
        print("Make sure the directory contains RST files with test results.")
        sys.exit(1)

    print(f"Found {len(test_data)} tests total")

    # Generate Excel report
    success = generate_excel_report(test_data, excel_path, args.version)

    if not success:
        print("Failed to generate Excel report")
        sys.exit(1)

    # Generate and print summary
    stats = get_test_summary_stats(test_data)
    print_summary_report(test_data, excel_path, args.version, stats)


if __name__ == "__main__":
    main()