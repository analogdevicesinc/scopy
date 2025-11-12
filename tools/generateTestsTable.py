#!/usr/bin/env python3
"""
Generate Excel file with all Scopy tests for execution tracking.
Supports version-based worksheets and RBP filtering/sorting.

Usage:
    python3 generateTestsTable.py v2.3.0                    # Create version worksheet
    python3 generateTestsTable.py v2.3.0 --sort             # Sort by RBP priority
    python3 generateTestsTable.py v2.3.0 --filter P0,P1     #Filter specific RBPs

Requirements:
    pip install openpyxl
"""

import os
import re
import argparse
import sys

# Import Excel dependencies
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl library is required for this script.")
    print("Install with: pip install openpyxl")
    sys.exit(1)

def extract_tests_from_rst(file_path):
    """Extract test UID and RBP pairs from an RST file."""
    tests = []

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Find all UID patterns
        uid_pattern = r'\*\*UID:\*\*\s+([A-Z_][A-Z0-9_.]*)'
        uid_matches = re.findall(uid_pattern, content)

        if not uid_matches:
            return tests

        # For each UID, find the corresponding RBP
        for uid in uid_matches:
            # Look for RBP after this UID
            uid_pos = content.find(f"**UID:** {uid}")
            if uid_pos == -1:
                continue

            # Search for RBP in the next 500 characters after UID
            search_area = content[uid_pos:uid_pos + 500]
            rbp_pattern = r'\*\*RBP:\*\*\s+(P[0-3])'
            rbp_match = re.search(rbp_pattern, search_area)

            if rbp_match:
                rbp = rbp_match.group(1)
                tests.append({
                    'uid': uid,
                    'rbp': rbp,
                    'file': os.path.basename(file_path)
                })
            else:
                # If no RBP found, report as missing
                tests.append({
                    'uid': uid,
                    'rbp': 'MISSING',
                    'file': os.path.basename(file_path)
                })

    except Exception as e:
        print(f"Error processing {file_path}: {e}")

    return tests

def get_all_tests():
    """Get all tests from RST files."""
    # Base directory for tests (relative to this script's location)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    scopy_root = os.path.dirname(script_dir)  # we assume script is in tools/
    test_base_dir = os.path.join(scopy_root, "docs", "tests")

    all_tests = []

    # Walk through all RST files
    for root, dirs, files in os.walk(test_base_dir):
        for file in files:
            if file.endswith('.rst'):
                file_path = os.path.join(root, file)
                tests = extract_tests_from_rst(file_path)
                all_tests.extend(tests)

    return all_tests

def filter_tests(tests, filter_rbp):
    """Filter tests by RBP levels."""
    if not filter_rbp:
        return tests

    filter_levels = [level.strip() for level in filter_rbp.split(',')]
    return [test for test in tests if test['rbp'] in filter_levels]

def sort_tests(tests, sort_by_rbp):
    """Sort tests by UID or RBP priority."""
    if sort_by_rbp:
        # Define RBP priority order
        rbp_order = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3, 'MISSING': 4}
        return sorted(tests, key=lambda x: (rbp_order.get(x['rbp'], 5), x['uid']))
 

def generate_excel_output(tests, excel_path, version_name):
    """Generate Excel file with version worksheet."""

    # Load existing workbook or create new one
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Create or get worksheet for this version
    if version_name in wb.sheetnames:
        ws = wb[version_name]
        # Clear existing content
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(version_name)

    row = 1
    # Write headers
    headers = ['Test UID', 'Test RBP', 'Test Pass/Fail']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
    row += 1
        
    # Write test data
    for test in tests:
        ws[f'A{row}'] = test['uid']
        ws[f'B{row}'] = test['rbp']
        ws[f'C{row}'] = ''  # Pass/Fail column
        
        row += 1

    # Auto-size columns
    for col in range(1, 4):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    # Save workbook
    wb.save(excel_path)
    return True

def print_summary(tests, excel_path, version_name):
    """Print generation summary."""
    # Generate summary statistics
    rbp_counts = {}
    missing_rbp = []

    for test in tests:
        if test['rbp'] == 'MISSING':
            missing_rbp.append(f"{test['uid']} (in {test['file']})")
        else:
            rbp_counts[test['rbp']] = rbp_counts.get(test['rbp'], 0) + 1

    # Print summary
    print(f"\n=== Scopy Test Generation Summary ===")
    print(f"Version: {version_name}")
    print(f"Total tests: {len(tests)}")
    print(f"Excel file: {excel_path}")
    print(f"\nRBP Distribution:")
    for rbp in ['P0', 'P1', 'P2', 'P3']:
        count = rbp_counts.get(rbp, 0)
        print(f"  {rbp}: {count} tests")

    if missing_rbp:
        print(f"\nTests missing RBP metadata ({len(missing_rbp)}):")
        for test in missing_rbp:
            print(f"  - {test}")
    else:
        print(f"\n✓ All tests have RBP metadata!")

def main():
    parser = argparse.ArgumentParser(
        description='Generate Scopy test tracking Excel file',
        epilog='Note: Requires openpyxl library (pip install openpyxl)'
    )
    parser.add_argument('version', help='Version name for Excel worksheet (e.g., v2.3.0)')
    parser.add_argument('--sort', action='store_true', help='Sort tests by RBP priority instead of UID')
    parser.add_argument('--filter', help='Filter tests by RBP levels (e.g., P0,P1)')

    args = parser.parse_args()

    # Version is required
    if not args.version:
        parser.error("Version name is required (e.g., v2.3.0)")

    print(f"Generating Excel worksheet for version: {args.version}")

    # Get all tests
    all_tests = get_all_tests()

    # Apply filters
    if args.filter:
        all_tests = filter_tests(all_tests, args.filter)
        print(f"Filtered to RBP levels: {args.filter}")

    # Apply sorting
    if args.sort:
        all_tests = sort_tests(all_tests, args.sort)
        print("Sorted by RBP priority")

    # Generate Excel file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    scopy_root = os.path.dirname(script_dir)
    excel_path = os.path.join(scopy_root, "scopy_tests_tracking.xlsx")

    success = generate_excel_output(all_tests, excel_path, args.version)

    if success:
        print_summary(all_tests, excel_path, args.version)
        print(f"\n✓ Excel worksheet '{args.version}' created successfully!")
    else:
        print("Failed to generate Excel file")
        sys.exit(1)

if __name__ == "__main__":
    main()